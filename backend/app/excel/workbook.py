from datetime import datetime
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.models.schemas import ProcessedOrder


def _timestamp() -> str:
    return datetime.now().strftime("%d%m%Y_%H%M%S")


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _find_header_cell(sheet, header: str) -> tuple[int, int] | None:
    normalized = header.strip().lower()
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 20)):
        for cell in row:
            if _normalize_header(cell.value) == normalized:
                return cell.row, cell.column
    return None


def _copy_cell(source, target) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format


def _catalog_bounds(sheet) -> tuple[int, int, int, int]:
    code_cell = _find_header_cell(sheet, "CODE")
    product_cell = _find_header_cell(sheet, "PRODUCT")
    offer_cell = _find_header_cell(sheet, "SPECIAL OFFER")
    if code_cell and product_cell and offer_cell:
        return code_cell[0], code_cell[1], offer_cell[1], product_cell[1]
    return 1, 1, sheet.max_column, 2


class WorkbookGenerator:
    def __init__(self, master_path: Path, downloads_dir: Path) -> None:
        self.master_path = master_path
        self.downloads_dir = downloads_dir
        self.downloads_dir.mkdir(parents=True, exist_ok=True)

    def generate_updated(self, order: ProcessedOrder) -> str:
        filename = f"Order_{_timestamp()}.xlsx"
        target = self.downloads_dir / filename
        if self.master_path.exists():
            workbook = load_workbook(self.master_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "ORDO Order"
            sheet.append(["CODE", "PRODUCT", "PACK SIZE", "PACK TYPE", "PTS", "PTR", "MRP", "Case", "DIVISION", "ORDER QTY", "ORDER VALUE", "SPECIAL OFFER"])
        fill = PatternFill(start_color="D9F99D", end_color="D9F99D", fill_type="solid")
        order_qty_cell = _find_header_cell(sheet, "ORDER QTY")
        if order_qty_cell:
            _, quantity_column = order_qty_cell
        else:
            quantity_column = sheet.max_column + 1
            sheet.cell(1, quantity_column).value = "Quantity Ordered"
        for row in order.rows:
            if not row.matchedProduct:
                continue
            row_number = row.matchedProduct.workbookRow
            if row_number <= sheet.max_row:
                sheet.cell(row_number, quantity_column).value = row.quantity
                for cell in sheet[row_number]:
                    cell.fill = fill
            else:
                product = row.matchedProduct
                next_row = sheet.max_row + 1
                sheet.append(["", product.name, product.pack, "", product.rate, "", "", "", product.company, row.quantity, f"=J{next_row}*E{next_row}", ""])
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill
        workbook.save(target)
        return filename

    def generate_items(self, order: ProcessedOrder) -> str:
        filename = f"Order_Items_{_timestamp()}.xlsx"
        target = self.downloads_dir / filename
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ordered Products"

        if self.master_path.exists():
            source_workbook = load_workbook(self.master_path, data_only=False)
            source_sheet = source_workbook.active
            header_row, start_col, end_col, _ = _catalog_bounds(source_sheet)
            order_qty_cell = _find_header_cell(source_sheet, "ORDER QTY")
            order_value_cell = _find_header_cell(source_sheet, "ORDER VALUE")
            quantity_column = order_qty_cell[1] - start_col + 1 if order_qty_cell else None
            value_column = order_value_cell[1] - start_col + 1 if order_value_cell else None
            pts_column = _find_header_cell(source_sheet, "PTS")
            pts_output_column = pts_column[1] - start_col + 1 if pts_column else None

            for col_index, source_col in enumerate(range(start_col, end_col + 1), start=1):
                target_cell = sheet.cell(1, col_index)
                _copy_cell(source_sheet.cell(header_row, source_col), target_cell)
                if isinstance(target_cell.value, str):
                    target_cell.value = target_cell.value.strip()
                source_letter = source_sheet.cell(header_row, source_col).column_letter
                sheet.column_dimensions[sheet.cell(1, col_index).column_letter].width = (
                    source_sheet.column_dimensions[source_letter].width or 14
                )

            output_row = 2
            for row in order.rows:
                if not row.matchedProduct:
                    continue
                source_row = row.matchedProduct.workbookRow
                for col_index, source_col in enumerate(range(start_col, end_col + 1), start=1):
                    _copy_cell(source_sheet.cell(source_row, source_col), sheet.cell(output_row, col_index))
                if quantity_column:
                    sheet.cell(output_row, quantity_column).value = row.quantity
                if value_column and quantity_column and pts_output_column:
                    sheet.cell(output_row, value_column).value = (
                        f"={sheet.cell(output_row, quantity_column).coordinate}"
                        f"*{sheet.cell(output_row, pts_output_column).coordinate}"
                    )
                output_row += 1
        else:
            sheet.append(["CODE", "PRODUCT", "PACK SIZE", "PACK TYPE", "PTS", "PTR", "MRP", "Case", "DIVISION", "ORDER QTY", "ORDER VALUE", "SPECIAL OFFER"])
            for row in order.rows:
                if not row.matchedProduct:
                    continue
                product = row.matchedProduct
                sheet.append(["", product.name, product.pack, "", product.rate, "", "", "", product.company, row.quantity, f"=J{sheet.max_row + 1}*E{sheet.max_row + 1}", ""])
        workbook.save(target)
        return filename
