import json
import re
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz import fuzz

from app.models.schemas import Product


PRODUCT_COLUMNS = {
    "name": ("product", "product name", "item", "item name", "medicine", "name"),
    "pack_size": ("pack size", "pack", "packing", "size"),
    "pack_type": ("pack type", "type"),
    "rate": ("pts", "rate", "price", "mrp"),
    "gst": ("gst", "tax"),
    "company": ("division", "company", "manufacturer", "mfr"),
}


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _search_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _find_columns(headers: list[str]) -> dict[str, int]:
    found: dict[str, int] = {}
    for field, aliases in PRODUCT_COLUMNS.items():
        for index, header in enumerate(headers):
            if header in aliases:
                found[field] = index
                break
    if "name" not in found:
        found["name"] = 0
    return found


def _find_header_row(sheet) -> tuple[int, dict[str, int]]:
    for row_index in range(1, min(sheet.max_row or 1, 20) + 1):
        row = next(sheet.iter_rows(min_row=row_index, max_row=row_index))
        headers = [_normalize_header(cell.value) for cell in row]
        columns = _find_columns(headers)
        if "name" in columns and headers[columns["name"]] in PRODUCT_COLUMNS["name"]:
            return row_index, columns
    row = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [_normalize_header(cell.value) for cell in row]
    return 1, _find_columns(headers)


def _coerce_float(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0


def _starter_catalog() -> list[Product]:
    bases = [
        "Azithral", "Dolo", "Pantocid", "Augmentin", "Cetrizine", "Metformin",
        "Telma", "Amlong", "Shelcal", "Zincovit", "Rantac", "Montair",
        "Pan", "Allegra", "Calpol", "Meftal", "Ecosprin", "Thyronorm",
        "Liv", "Becosules", "Clavam", "Taxim", "Omez", "Rosuvas",
        "Atorva", "Glycomet", "Istamet", "Janumet", "Telmikind", "Olmezest",
        "Aciloc", "Nexpro", "Montek", "Deriphyllin", "Sinarest", "Combiflam",
        "Emeset", "Ondem", "Cyclopam", "Sporlac",
    ]
    strengths = ["50", "75", "100", "120", "150", "250", "500", "625", "650", "DS", "SR", "Duo"]
    packs = ["10 TAB", "15 TAB", "10 CAP", "30 TAB", "100 ML", "60 ML", "1 VIAL"]
    companies = ["Cipla", "Sun Pharma", "Abbott", "Alkem", "Glenmark", "Mankind", "Dr Reddy", "Torrent"]
    products: list[Product] = []
    for index in range(1, 647):
        base = bases[(index - 1) % len(bases)]
        strength = strengths[((index - 1) // len(bases)) % len(strengths)]
        name = f"{base} {strength}"
        products.append(
            Product(
                id=f"{_slug(name)}-{index}",
                name=name,
                pack=packs[index % len(packs)],
                rate=round(32 + index * 4.7, 2),
                gst=5 if index % 5 == 0 else 12,
                company=companies[index % len(companies)],
                workbookRow=index + 1,
            )
        )
    return products


class ProductCatalog:
    def __init__(self, workbook_path: Path, cache_path: Path) -> None:
        self.workbook_path = workbook_path
        self.cache_path = cache_path
        self.products: list[Product] = []
        self.source = "starter catalog"

    def load(self) -> None:
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        if self.workbook_path.exists():
            self.products = self._load_from_workbook()
            self.source = str(self.workbook_path)
            self._write_cache()
            return
        if self.cache_path.exists():
            self.products = [Product.model_validate(item) for item in json.loads(self.cache_path.read_text())]
            self.source = str(self.cache_path)
            return
        self.products = _starter_catalog()
        self._write_cache()

    def _load_from_workbook(self) -> list[Product]:
        workbook = load_workbook(self.workbook_path, data_only=False)
        sheet = workbook.active
        header_row, columns = _find_header_row(sheet)
        products: list[Product] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            name = str(row[columns["name"]].value or "").strip()
            if not name:
                continue
            pack_parts = [
                str(row[columns[key]].value or "").strip()
                for key in ("pack_size", "pack_type")
                if key in columns and str(row[columns[key]].value or "").strip()
            ]
            pack = " ".join(pack_parts)
            company = str(row[columns.get("company", columns["name"])].value or "").strip()
            products.append(
                Product(
                    id=f"{_slug(name)}-{row_index}",
                    name=name,
                    pack=pack or "-",
                    rate=_coerce_float(row[columns.get("rate", columns["name"])].value),
                    gst=_coerce_float(row[columns["gst"]].value) if "gst" in columns else 0,
                    company=company or "-",
                    workbookRow=row_index,
                )
            )
        return products

    def _write_cache(self) -> None:
        self.cache_path.write_text(
            json.dumps([product.model_dump() for product in self.products], indent=2),
            encoding="utf-8",
        )

    def search(self, query: str, limit: int = 80) -> list[Product]:
        normalized = query.strip().lower()
        if not normalized:
            return self.products[:limit]
        query_key = _search_key(normalized)
        ranked: list[tuple[float, Product]] = []
        for product in self.products:
            haystack_key = _search_key(f"{product.name} {product.company} {product.pack}")
            score = max(
                100 if query_key in haystack_key else 0,
                fuzz.WRatio(query_key, haystack_key),
                fuzz.partial_ratio(query_key, haystack_key),
            )
            if score >= 45:
                ranked.append((float(score), product))
        return [product for _, product in sorted(ranked, key=lambda item: (-item[0], item[1].name))[:limit]]
